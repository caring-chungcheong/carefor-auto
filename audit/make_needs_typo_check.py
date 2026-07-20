# -*- coding: utf-8 -*-
"""욕구사정 판단근거·총평 오타 검수 (로컬 전용).

⚠️ 한글 맞춤법 전수 검사는 외부 API(네이버 등)가 필요한데, 그건 수급자 의료정보를
   제3자에게 보내는 것 → 개인정보 원칙 위반. **로컬로는 절대 안 함.**
   대신 (1) **명백히 틀린 상용오타 사전**(문맥 무관하게 항상 틀린 형태만) +
        (2) **깨진 낱자**(완성 안 된 한글 자모) 만 본다. 오탐 0 을 목표로 보수적으로.
   실측: 청주 판단근거 73만자에서 '뇌졸증'(→뇌졸중) 1건 정확 검출, 오탐 없음.

되/돼, 안/않, 로서/로써, 든지/던지, 바램(바람) 등 **문맥에 따라 맞을 수 있는 것은 제외**
(넣으면 오탐 홍수). 새 오타는 TYPOS 에 '틀린형태: 바른표기' 로만 추가하면 된다.

실행: py -X utf8 -m audit.make_needs_typo_check [지점키=청주]
출력: 평가준비/<지점>/<지점>_오타검수.xlsx  (개인정보 → 로컬만, repo 밖)
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from audit.deskpath import out_dir

RES = Path(__file__).resolve().parent.parent / "audit_results"

# 명백히 틀린 형태만(어떤 문맥에서도 이 표기는 오타). 값 = 바른 표기.
# ★★ 절대 규칙: **정상 단어의 일부(substring)가 되는 형태는 넣지 말 것.**
#   실측 사고: "치매증"→"치매" 를 넣었더니 올바른 "치매증**상**" 73건을 오타로 오탐.
#   ("우울증증"⊂우울증증상 도 같은 문제). → 그 자체로 완결된 오타 형태만 등록한다.
TYPOS = {
    # 의료·돌봄 (전체가 오타인 형태만 — '치매증/우울증증' 류는 정상단어 부분이라 제외)
    "뇌졸증": "뇌졸중", "골다골증": "골다공증", "파키슨": "파킨슨", "파킨스": "파킨슨",
    "고혈합": "고혈압", "관절렴": "관절염",
    # 일반 상용(문맥 무관하게 틀린 것만. 되/돼·안/않·바램 등 문맥 애매한 것은 제외)
    "몇일": "며칠", "희안": "희한", "금새": "금세", "역활": "역할", "오랫만": "오랜만",
    "어떻해": "어떡해", "폭팔": "폭발", "통채": "통째", "닥달": "닦달", "서슴치": "서슴지",
    "베게": "베개", "설겆이": "설거지", "왠만": "웬만",
    "할께": "할게", "갈께": "갈게", "먹을께": "먹을게", "됬": "됐", "안됫": "안됐",
    "바꼇": "바뀌었", "내노라": "내로라", "설레임": "설렘", "메세지": "메시지",
}
# 완성 안 된 한글 낱자(자모 단독) — 깨진 글자
JAMO = re.compile(r"[ㄱ-ㅎㅏ-ㅣ]")


def find_typos(text):
    """(유형, 발견, 바른표기, 문맥) 리스트."""
    t = (text or "").replace("\u200E", "")
    out = []
    for wrong, right in TYPOS.items():
        for m in re.finditer(re.escape(wrong), t):
            ctx = t[max(0, m.start() - 12):m.end() + 12].replace("\n", " ")
            out.append(("오타", wrong, right, ctx))
    for m in JAMO.finditer(t):
        ctx = t[max(0, m.start() - 12):m.end() + 12].replace("\n", " ")
        out.append(("깨진 낱자", m.group(0), "완성 글자 확인", ctx))
    return out


def build(branch_key):
    src = next(RES.glob(f"needs_full_*{branch_key}*.json"))
    d = json.loads(src.read_text(encoding="utf-8"))
    cutoff = d.get("cutoff", "")
    rows = []
    for p in d["people"]:
        for a in p.get("assess", []):
            dt = a.get("date", "")
            if cutoff and dt and dt < cutoff:
                continue
            for r in a.get("rows", []):
                lb = str(r.get("label", ""))
                if "판단근거" not in lb and "총평" not in lb:
                    continue
                for kind, found, right, ctx in find_typos(r.get("text", "")):
                    rows.append([p["name"], dt, lb, kind, found, right, ctx])
    return rows, cutoff


def main():
    key = sys.argv[1] if len(sys.argv) > 1 else "청주"
    data, cutoff = build(key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "오타 검수"
    ws.append(["수급자", "사정일", "항목", "유형", "발견", "바른 표기", "앞뒤 문맥"])
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="A94464")
    thin = Side(style="thin", color="E0D3D8")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
    warn = PatternFill("solid", fgColor="FFF2CC")
    for r in sorted(data, key=lambda x: (str(x[0]), str(x[1]))):
        ws.append(r)
        rr = ws.max_row
        for c in ws[rr]:
            c.border = bd
            c.alignment = Alignment(vertical="center", wrap_text=(c.column == 7))
            if c.column in (5, 6):
                c.fill = warn
    for col, w in zip("ABCDEFG", [10, 12, 20, 10, 12, 12, 44]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    out = out_dir(key) / f"{key}_오타검수.xlsx"
    wb.save(out)
    print(f"저장: {out}")
    print(f"오타·깨진글자 {len(data)}건 · 평가기간 {cutoff}~")
    print("※ 명백한 상용오타 + 깨진 낱자만(문맥 애매한 되/돼·안/않 제외). 새 오타는 TYPOS에 추가.")


if __name__ == "__main__":
    main()
