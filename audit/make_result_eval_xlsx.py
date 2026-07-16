# -*- coding: utf-8 -*-
"""결과평가 다운로드분(audit_results/결과평가_<지점>/)을 엑셀 정리본으로 생성.

- HTML의 <g-th>/<g-td> 쌍에서 인적사항·체크사항을, case_tit 구조에서 총평을 파싱.
- 수급자별 폴더 구조 그대로 산출 폴더에 복사하고 엑셀에서 하이퍼링크.
- '명단점검' 시트: 리스트 상단 인원수 vs 수집 명단 대조 + 인원별 연도별 건수.

실행: py -X utf8 -m audit.make_result_eval_xlsx [지점키=청주]
결과: 클로드코드/<지점>_급여제공결과평가.xlsx + <지점>_급여제공결과평가_원본/
"""
from __future__ import annotations
import sys, json, re, shutil
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent          # carefor-auto
RES = BASE / "audit_results"
OUT = BASE.parent                                       # 클로드코드

TAG = re.compile(r"<[^>]+>")
BR = re.compile(r"<br\s*/?>", re.I)
PAIR = re.compile(r"<g-th[^>]*>(.*?)</g-th>\s*<g-td[^>]*>(.*?)</g-td>", re.S)
TOTAL = re.compile(r'case_tit[^>]*>\s*총평\s*</div>.*?<g-td[^>]*>(.*?)</g-td>', re.S)

F = dict(name="맑은 고딕", size=10)
FILL = PatternFill("solid", fgColor="4472C4")
WARN = PatternFill("solid", fgColor="FCE4EC")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", TAG.sub(" ", s)).strip()


def multiline(s: str) -> str:
    s = BR.sub("\n", s)
    lines = [re.sub(r"\s+", " ", TAG.sub(" ", ln)).strip() for ln in s.split("\n")]
    return "\n".join(lines).strip()


def parse_eval(html: str) -> dict:
    d = {"생년월일": "", "성별": "", "급여개시일": "", "재급여개시일": "", "퇴소일": "",
         "케어그룹": "", "작성일": "", "작성자": "", "c1": "", "c2": "", "c3": "", "c4": "", "총평": ""}
    for th_raw, td_raw in PAIR.findall(html):
        th, td = clean(th_raw), clean(td_raw)
        if th.startswith("생년월일"):
            m = re.search(r"\d{4}\.\d{2}\.\d{2}", td)
            d["생년월일"] = m.group(0) if m else td
        elif th.startswith("성별"):
            d["성별"] = td
        elif th.startswith("급여개시일"):
            m = re.search(r"(\d{4}\.\d{2}\.\d{2})", td)
            d["급여개시일"] = m.group(1) if m else td
            m2 = re.search(r"재급여개시일\s*:\s*(\d{4}\.\d{2}\.\d{2})", td)
            if m2:
                d["재급여개시일"] = m2.group(1)
        elif th.startswith("퇴소일"):
            d["퇴소일"] = td
        elif th.startswith("케어그룹"):
            d["케어그룹"] = td
        elif th.startswith("작성일"):
            d["작성일"] = td
        elif th.startswith("작성자"):
            d["작성자"] = td
        elif th.startswith("1."):
            d["c1"] = td
        elif th.startswith("2."):
            d["c2"] = td
        elif th.startswith("3."):
            d["c3"] = td
        elif th.startswith("4."):
            d["c4"] = td
    m = TOTAL.search(html)
    if m:
        d["총평"] = multiline(m.group(1))
    return d


def style_header(ws, ncol):
    for c in ws[ws.max_row][:ncol]:
        c.font = Font(**F, bold=True, color="FFFFFF")
        c.fill, c.border = FILL, THIN
        c.alignment = Alignment(horizontal="center", vertical="center")


HEAD = ["수급자명", "현황", "생년월일", "성별", "급여개시일", "재급여개시일", "퇴소일", "케어그룹",
        "작성일(평가일)", "작성자", "1.진행된 급여서비스", "2.욕구반영 여부", "3.상태변화",
        "4.계획 재작성 필요", "총평", "원본파일"]
WIDTH = [10, 8, 11, 6, 11, 12, 11, 9, 13, 9, 30, 12, 10, 16, 90, 38]


def main():
    key = sys.argv[1] if len(sys.argv) > 1 else "청주"
    src = next(p for p in RES.glob(f"결과평가_*{key}*") if p.is_dir())
    branch = src.name.replace("결과평가_", "")
    idx = json.loads((src / "index.json").read_text(encoding="utf-8"))
    header, check, roster, evals = idx.get("header", {}), idx.get("check", {}), idx.get("roster", []), idx["evals"]

    # 원본 HTML 을 수급자별 폴더 구조 그대로 복사
    raw_dir = OUT / f"{branch}_급여제공결과평가_원본"
    if raw_dir.exists():
        shutil.rmtree(raw_dir)
    shutil.copytree(src, raw_dir, ignore=shutil.ignore_patterns("index.json"))

    rows = []
    for e in evals:
        d = parse_eval((src / e["file"]).read_text(encoding="utf-8"))
        rows.append({**e, **d})
    rows.sort(key=lambda r: (r["name"], r["작성일"] or r["date"]))

    wb = Workbook()

    # ── 시트1: 명단점검 ──
    ws = wb.active
    ws.title = "명단점검"
    per = {}
    for r in rows:
        p = per.setdefault(r["pammgno"], {"cnt": 0, "y": {}, "last": ""})
        p["cnt"] += 1
        yy = (r["작성일"] or r["date"])[:4]
        p["y"][yy] = p["y"].get(yy, 0) + 1
        p["last"] = max(p["last"], r["작성일"] or r["date"])
    ws.append(["구분", "인원수", "비고"])
    style_header(ws, 3)
    summary = [
        ("케어포 리스트 상단 표시(전체)", header.get("total"), f"남 {header.get('male','?')} / 여 {header.get('female','?')}"),
        ("다운로드한 명단(퇴소자 포함)", check.get("목록_인원"), ""),
        ("스캔한 인원", check.get("스캔_인원"), ""),
        ("결과평가 보유 인원", check.get("평가보유_인원"), "2024년 이후 작성분 기준"),
        ("판정", None, "일치" if check.get("일치") else "불일치 — 미스캔 인원 확인 필요"),
    ]
    for a, b_, c in summary:
        ws.append([a, b_, c])
        for cell in ws[ws.max_row][:3]:
            cell.font, cell.border = Font(**F), THIN
        if a == "판정" and not check.get("일치"):
            for cell in ws[ws.max_row][:3]:
                cell.fill = WARN
    ws.append([])
    ws.append(["수급자명", "현황", "평가건수", "2024", "2025", "2026", "최근 작성일", "폴더"])
    style_header(ws, 8)
    for p in roster:
        g = per.get(p["pammgno"], {"cnt": 0, "y": {}, "last": ""})
        folder = re.sub(r'[\\/:*?"<>|]', "_", f"{p['name']}_{p['pammgno']}")
        ws.append([p["name"], p["status"], g["cnt"], g["y"].get("2024", 0), g["y"].get("2025", 0),
                   g["y"].get("2026", 0), g["last"], folder if g["cnt"] else ""])
        row = ws.max_row
        for cell in ws[row][:8]:
            cell.font, cell.border = Font(**F), THIN
        if g["cnt"] == 0:
            for cell in ws[row][:8]:
                cell.fill = WARN
        else:
            link = ws.cell(row=row, column=8)
            link.hyperlink = f"{branch}_급여제공결과평가_원본/{folder}"
            link.font = Font(**F, color="0563C1", underline="single")
    for i, w in enumerate([26, 8, 9, 6, 6, 6, 12, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:H{ws.max_row}"

    # ── 시트2: 결과평가 전체 목록 ──
    ws = wb.create_sheet("결과평가")
    ws.append(HEAD)
    style_header(ws, len(HEAD))
    for r in rows:
        ws.append([r["name"], r["status"], r["생년월일"], r["성별"], r["급여개시일"], r["재급여개시일"],
                   r["퇴소일"], r["케어그룹"], r["작성일"] or r["date"], r["작성자"],
                   r["c1"], r["c2"], r["c3"], r["c4"], r["총평"], r["file"]])
        row = ws.max_row
        for c in ws[row]:
            c.font, c.border = Font(**F), THIN
            c.alignment = Alignment(vertical="top", wrap_text=c.column in (11, 15))
        link = ws.cell(row=row, column=16)
        link.hyperlink = f"{branch}_급여제공결과평가_원본/{r['file']}"
        link.font = Font(**F, color="0563C1", underline="single")
    for i, w in enumerate(WIDTH, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{ws.max_row}"

    out = OUT / f"{branch}_급여제공결과평가.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = OUT / f"{branch}_급여제공결과평가_수정본.xlsx"
        wb.save(out)
    print(f"저장: {out} ({len(rows)}건) / 원본 HTML: {raw_dir}")
    miss = [r["file"] for r in rows if not r["작성일"] or not r["총평"]]
    if miss:
        print(f"※ 작성일/총평 파싱 공란 {len(miss)}건: {miss[:5]}")


if __name__ == "__main__":
    main()
