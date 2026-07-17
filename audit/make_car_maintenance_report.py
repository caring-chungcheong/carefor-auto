# -*- coding: utf-8 -*-
"""차량 정비이력 한눈에 보기 + 수리비 입력·월별 집계 + 타이어 사이즈 마스터.

사전: py -X utf8 -m audit.collect_car_maintenance <지점> --from <개소일/cutoff>
실행: py -X utf8 -m audit.make_car_maintenance_report <지점>
결과: 클로드코드/차량정비이력/<지점>/ — 엑셀 + 차량별 첨부 PDF

시트 구성:
  · 정비이력   — 목록 + **수리비(원) 입력칸** + 첨부 PDF 하이퍼링크 (타이어 교체는 노랑)
  · 월별 수리비 — 정비이력의 수리비를 SUMIFS 로 자동 집계 (입력하면 즉시 갱신)
  · 차량마스터 — 차량번호·차종·연식 + **타이어 사이즈 입력칸** (차량별 1회 조사)

⚠️ 왜 수리비를 손으로 넣나: 케어포에 금액 필드가 아예 없고(그리드·내역서 폼 모두),
   금액은 첨부 PDF 안에만 있는데 그 PDF 는 정비소 영수증을 찍은 **스캔 이미지**라
   (폰트 0·JPEG·추출텍스트 0) 자동 추출이 불가하다. OCR 은 금액 오독 위험이 커 쓰지 않는다.
   → 첨부 링크를 눌러 PDF 를 보고 금액만 적으면 월별 집계가 자동으로 채워지게 했다.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit.ocr_car_invoice import find_sizes


def sizes_in(text: str) -> list[str]:
    """텍스트에서 유효한 타이어 규격 추출.
    ⚠️ 사이즈는 **두 곳**에 있다 — 첨부 PDF(OCR) 와 **케어포 정비내역 텍스트**.
       실측: 서구 아반떼2741 '225/45R17,04,H318 타이어교체' 는 첨부가 아예 없고 내역에만 있다.
       PDF 만 보면 통째로 놓친다."""
    return find_sizes(text)

sys.stdout.reconfigure(encoding="utf-8")
RES = Path(__file__).resolve().parent.parent / "audit_results"
DESK = Path(r"C:\Users\alsgm\OneDrive\Desktop\클로드코드\차량정비이력")

HDR_FILL = PatternFill("solid", fgColor="305496")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")   # 입력칸(노랑)
TIRE_FILL = PatternFill("solid", fgColor="FCE4D6")  # 타이어 교체 행(주황)

# 케어포 정비구분은 고장수리/예방정비/정기점검/기타뿐 → 부품은 정비내역 자유텍스트에서 키워드로
KINDS = [
    ("타이어", re.compile(r"타이어")),
    ("브레이크", re.compile(r"브레이크|라이닝|패드|디스크")),
    ("배터리", re.compile(r"배터리")),
    ("엔진오일", re.compile(r"엔진\s*오일|엔진오일")),
]
# 분기 정기점검 상용구 — '점검했다'는 기록이지 교체가 아님 (타이어 334건 중 242건이 이것)
BOILER = re.compile(r"마모상태\s*점검")


def classify(desc: str) -> tuple[str, str]:
    hits = [n for n, p in KINDS if p.search(desc)]
    if not hits:
        return "", ""
    return "·".join(hits), ("점검" if BOILER.search(desc) else "교체/수리")


def _style_header(ws, widths):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("branch", nargs="?", default="청주")
    args = ap.parse_args()

    src = next((d for d in RES.glob("정비이력_*") if args.branch in d.name), None)
    if not src:
        print(f"수집물 없음. 먼저: py -X utf8 -m audit.collect_car_maintenance {args.branch}")
        sys.exit(1)
    data = json.loads((src / "index.json").read_text(encoding="utf-8"))
    bname, recs, cars = data["branch"], data["records"], data.get("cars", [])

    out = DESK / bname
    out.mkdir(parents=True, exist_ok=True)
    for d in src.iterdir():
        if d.is_dir():
            shutil.copytree(d, out / d.name, dirs_exist_ok=True)

    wb = openpyxl.Workbook()

    # ── 1) 정비이력 ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "정비이력"
    ws.append(["차량", "차종", "정비일", "연월", "정비구분", "부품", "구분",
               "정비내역", "첨부(내역서)", "OCR 금액(참고)", "OCR 사이즈", "확인필요",
               "수리비(원) 확정"])
    _style_header(ws, [14, 10, 11, 8, 9, 13, 9, 44, 40, 14, 13, 22, 14])
    for r in sorted(recs, key=lambda x: (x["car"], x["date"]), reverse=True):
        tag, mode = classify(r["desc"])
        f = r["files"][0] if r["files"] else ""
        ym = r["date"][:7].replace(".", "-") if r.get("date") else ""
        o = r.get("ocr") or {}
        warn = []
        if o:
            if not o.get("total"):
                warn.append("금액 못읽음")
            if o.get("pages", 1) > 1:
                warn.append(f"{o['pages']}페이지(총액 뒷장 가능)")
            if o.get("insurance"):
                warn.append("보험수리(지점부담 아닐 수 있음)")
        ws.append([r["car"], r["kind"], r["date"], ym, r["type"], tag, mode,
                   r["desc"], f or "없음",
                   o.get("total"), o.get("size") or (sizes_in(r["desc"])[0] if sizes_in(r["desc"]) else ""),
                   " / ".join(warn), None])
        row = ws.max_row
        if f:
            c = ws.cell(row, 9)
            c.hyperlink = f"{r['car']}/{f}"
            c.font = Font(color="0563C1", underline="single")
        ws.cell(row, 10).number_format = "#,##0"      # OCR 참고값
        ws.cell(row, 13).fill = IN_FILL               # 확정 수리비 입력칸
        ws.cell(row, 13).number_format = "#,##0"
        if warn:
            ws.cell(row, 12).font = Font(color="C00000")
        if tag and "타이어" in tag and mode == "교체/수리":
            for i in range(1, 13):
                ws.cell(row, i).fill = TIRE_FILL
    n = ws.max_row

    # ── 2) 월별 수리비 (SUMIFS — 수리비 입력하면 자동 갱신) ──────────
    ws2 = wb.create_sheet("월별 수리비")
    yms = sorted({r["date"][:7].replace(".", "-") for r in recs if r.get("date")})
    carnames = sorted({r["car"] for r in recs})
    ws2.append(["연월"] + carnames + ["합계(확정)", "건수(확정입력)", "OCR 참고합계"])
    _style_header(ws2, [10] + [14] * len(carnames) + [14, 15, 15])
    for ym in yms:
        row = [ym]
        for cn in carnames:
            row.append(f'=SUMIFS(정비이력!$M:$M,정비이력!$D:$D,$A{ws2.max_row+1},정비이력!$A:$A,'
                       f'"{cn}")')
        row.append(f'=SUM(B{ws2.max_row+1}:{get_column_letter(1+len(carnames))}{ws2.max_row+1})')
        row.append(f'=COUNTIFS(정비이력!$D:$D,$A{ws2.max_row+1},정비이력!$M:$M,">0")')
        row.append(f'=SUMIFS(정비이력!$J:$J,정비이력!$D:$D,$A{ws2.max_row+1})')
        ws2.append(row)
        for i in range(2, len(carnames) + 5):
            ws2.cell(ws2.max_row, i).number_format = "#,##0"
    last = ws2.max_row
    ws2.append(["총계"] + [f'=SUM({get_column_letter(i)}2:{get_column_letter(i)}{last})'
                          for i in range(2, len(carnames) + 5)])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)
        c.number_format = "#,##0"

    # ── 3) 차량마스터 (타이어 사이즈 1회 조사) ────────────────────
    ws3 = wb.create_sheet("차량마스터")
    ws3.append(["차량명", "차량번호", "차종", "연식", "운행시작", "타이어 사이즈", "근거"])
    _style_header(ws3, [16, 14, 12, 8, 12, 16, 56])
    # 정비내역 텍스트에서 나온 사이즈도 합친다 (첨부 PDF 없는 차량 구제)
    from collections import defaultdict
    desc_sz = defaultdict(set)
    for r in recs:
        for z in sizes_in(r["desc"]):
            desc_sz[r["car"]].add((z, r["date"]))
    for c in cars:
        st = c.get("stdt") or ""
        st = f"{st[:4]}.{st[4:6]}.{st[6:8]}" if len(st) == 8 else st
        sz, src = c.get("tire_size"), c.get("tire_src", "")
        if not sz and c.get("name") in desc_sz:
            z, dt = sorted(desc_sz[c["name"]])[0]
            sz, src = z, f"케어포 정비내역 텍스트 {dt} (첨부 없음)"
        ws3.append([c.get("name", ""), c.get("numb", ""), c.get("kind", ""),
                    c.get("modl", ""), st, sz or None, src])
        if not sz:
            ws3.cell(ws3.max_row, 6).fill = IN_FILL      # 미확인 → 입력 필요
            ws3.cell(ws3.max_row, 7).value = "명세서·정비내역에 사이즈 기재 없음 — 차량에서 직접 확인 필요"

    # ── 4) 안내 ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("안내")
    for line in [
        ["차량 정비이력 — 사용법"],
        [""],
        ["1) 정비이력 시트의 '첨부(내역서)' 링크를 누르면 정비소 영수증 PDF가 열립니다."],
        ["2) 그 PDF를 보고 '수리비(원)' 칸(노란색)에 금액만 적으면"],
        ["   '월별 수리비' 시트가 자동으로 합계를 냅니다 (수식이라 저장만 하면 갱신)."],
        ["3) '차량마스터' 시트의 '타이어 사이즈' 칸은 차량별로 1회만 조사해 넣으면 됩니다."],
        [""],
        ["왜 수리비를 손으로 넣나?"],
        ["  · 케어포에 금액 필드가 아예 없습니다(정비기록 그리드·내역서 폼 모두)."],
        ["  · 금액은 첨부 PDF 안에만 있는데, 그 PDF는 영수증을 찍은 스캔 이미지라"],
        ["    (폰트 없음·추출 텍스트 0) 자동으로 읽을 수 없습니다."],
        ["  · OCR은 금액을 잘못 읽을 위험이 커서 쓰지 않았습니다 — 틀린 집계가 더 위험합니다."],
        [""],
        ["왜 타이어 사이즈를 손으로 넣나?"],
        ["  · 케어포 정비내역 554건 중 사이즈가 적힌 건 단 1건입니다(자유텍스트라 입력 규칙이 없음)."],
        ["  · 앞으로도 안 쌓이므로 차량번호별 고정값으로 관리하는 편이 정확합니다."],
        [""],
        ["참고 — '부품'과 '구분'"],
        ["  · 케어포 정비구분은 고장수리/예방정비/정기점검/기타뿐이라, 타이어·브레이크·배터리는"],
        ["    정비내역 텍스트에서 키워드로 뽑았습니다."],
        ["  · '점검'은 분기마다 붙는 '타이어 마모상태 점검' 상용구이고, 실제 교체는 '교체/수리'입니다."],
        ["    (타이어 334건 중 242건이 상용구 — 그냥 세면 실제 교체가 묻힙니다)"],
    ]:
        ws4.append(line)
    ws4.column_dimensions["A"].width = 95
    ws4["A1"].font = Font(bold=True, size=13)

    fp = out / f"차량정비이력_{bname}.xlsx"
    wb.save(fp)
    att = sum(1 for r in recs if r["files"])
    print(f"저장: {fp}")
    print(f"  정비 {len(recs)}건 · 첨부 {att}건 · 차량 {len(cars)}대 · 월 {len(yms)}개월")


if __name__ == "__main__":
    main()
