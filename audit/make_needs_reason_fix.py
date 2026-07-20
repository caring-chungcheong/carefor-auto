# -*- coding: utf-8 -*-
"""욕구사정 '판단근거 미언급' 수정 지시서 — 체크는 했는데 판단근거를 안 쓴 건만.

make_needs_check_xlsx 산출(욕구사정_체크누락_<지점>.xlsx)의 '문제요약'을 파싱해
수급자·사정일·항목·체크값별로 **무엇을 어떻게 써야 하는지**를 한 줄씩 지시.
사례집(2026 평가대비 20번 최고위험): "병명만 적고 사유 없음 → 진단명으로 인한 → 구체적 증상 →
어떤 도움 필요" 형식. 이건 make_needs_fix_xlsx(낙상/욕창 대조)와 **다른** 체크다.

실행: py -X utf8 -m audit.make_needs_reason_fix [지점키=청주]
출력: 평가준비/<지점>/<지점>_판단근거_수정지시서.xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from audit.deskpath import out_dir

RES = Path(__file__).resolve().parent.parent / "audit_results"

# 항목 → 판단근거 작성 예시 템플릿(사례집 형식). {체크값} 채워 안내.
GUIDE = {
    "영양": "‘{v}’ 상태의 원인·정도와 그에 따른 도움 필요 여부를 서술 (예: OOO로 인해 {v} 있어 …도움 필요)",
    "질병": "‘{v}’ 진단으로 인한 구체적 증상과 일상 영향·도움을 서술 (예: {v}(으)로 인한 OO 증상으로 …)",
    "구강": "‘{v}’ 상태와 그로 인한 식사·구강관리 도움 필요를 서술 (예: {v}(으)로 저작곤란하여 …)",
    "가족": "‘{v}’ 동거·수발 관계와 지원 정도를 서술 (예: {v}와(과) 동거하며 주수발 …)",
    "구강(치아상태": "‘{v}’ 상태로 인한 저작·식사 도움 필요를 서술",
}
NEUTRAL = "체크한 ‘{v}’에 대한 판단근거(원인·정도·필요 도움)를 항목 서술란에 기재"
BODY_GUIDE = "신체상태 ‘{v}’ 항목의 완전자립/부분도움/완전도움 판단근거를 서술 (왜 그 도움이 필요한지)"


def parse_summary(s: str):
    """문제요약 → [(유형, 항목, 세부, 체크값), ...]

    형식이 중첩·복합이라(예: '영양(배설 양상 / 소변상태=요실금, 대변상태=자주 실수하심) / 구강(...)')
    전역 findall 로 '항목(세부=값)' 을 모두 뽑는다. 못 뽑은 잔여는 통째로 한 줄 남겨 놓친 게 없게 한다.
    """
    s = s or ""
    out = []
    # 신체 판단근거 미서술: A, B, C
    for m2 in re.finditer(r"신체 판단근거 미서술:\s*([^/]+)", s):
        for act in re.split(r"\s*,\s*", m2.group(1)):
            if act.strip():
                out.append(("신체서술", "신체상태", act.strip(), ""))
    # 항목(...=값) 을 전부 (괄호 안에 다시 괄호는 없다고 보고 [^()])
    for m in re.finditer(r"([가-힣]+)\(([^()]*?)=([^()]*?)\)", s):
        field, detail, val = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
        # detail 에 '소변상태' 처럼 마지막 키가 있으면 그걸 세부로
        detail = detail.split("/")[-1].strip() or detail
        out.append(("판단근거", field, detail, val))
    # 아무것도 못 뽑았는데 '미언급' 은 있으면 원문을 통째로
    if not out and "미언급" in s:
        raw = re.sub(r".*미언급\s*[—\-]\s*", "", s).strip()
        out.append(("판단근거", raw or s, "", ""))
    return out


def guide_for(kind, field, detail, val):
    if kind == "신체서술":
        return BODY_GUIDE.format(v=detail)
    tmpl = GUIDE.get(field) or GUIDE.get(field.split("(")[0]) or NEUTRAL
    return tmpl.format(v=(val or detail or "해당 항목"))


def main():
    key = sys.argv[1] if len(sys.argv) > 1 else "청주"
    src = next(RES.glob(f"욕구사정_체크누락_*{key}*.xlsx"), None)
    if not src:
        # 체크표가 없으면 만들라고 안내
        src2 = next((out_dir(key)).glob(f"욕구사정_체크누락_*.xlsx"), None)
        src = src2
    if not src:
        print(f"체크표 없음. 먼저: py -X utf8 -m audit.make_needs_check_xlsx {key}")
        sys.exit(1)

    wb0 = openpyxl.load_workbook(src)
    # 체크표는 시트 3개(점검결과·문제목록·판단근거_미언급). '문제목록'이 문제건만 깔끔히 담는다.
    ws0 = wb0["문제목록"] if "문제목록" in wb0.sheetnames else wb0.active
    hdr = [c.value for c in ws0[1]]
    ix = {h: i for i, h in enumerate(hdr) if h}
    ci_name, ci_bd, ci_dt = ix["수급자"], ix["생년월일"], ix["사정일"]
    ci_s = ix.get("문제") if "문제" in ix else ix.get("문제요약")

    rows = []
    for r in ws0.iter_rows(min_row=2):
        v = [c.value for c in r]
        summ = v[ci_s]
        if not summ or not str(summ).strip():
            continue
        for kind, field, detail, val in parse_summary(summ):
            label = f"{field}({detail}={val})" if val else (f"{field}: {detail}" if detail else field)
            rows.append([v[ci_name], v[ci_bd], v[ci_dt], field, label, guide_for(kind, field, detail, val)])

    # 엑셀 작성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판단근거 수정지시"
    heads = ["수급자", "생년월일", "사정일", "항목", "현재 체크(근거 없음)", "무엇을 써야 하나", "완료"]
    ws.append(heads)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="A94464")
    thin = Side(style="thin", color="E0D3D8")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
    prev = None
    for row in sorted(rows, key=lambda x: (str(x[0]), str(x[2]))):
        ws.append(row)
        rr = ws.max_row
        # 사람 바뀌는 경계 옅은 음영
        band = PatternFill("solid", fgColor="FBF5F7") if (row[0] != prev) else None
        prev = row[0]
        for c in ws[rr]:
            c.border = bd
            c.alignment = Alignment(vertical="center", wrap_text=(c.column in (5, 6)))
            if band:
                c.fill = band
    widths = [10, 12, 12, 10, 34, 52, 6]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = out_dir(key) / f"{key}_판단근거_수정지시서.xlsx"
    wb.save(out)
    npeople = len({r[0] for r in rows})
    print(f"저장: {out}")
    print(f"수정 대상 {npeople}명 · 항목 {len(rows)}건")


if __name__ == "__main__":
    main()
