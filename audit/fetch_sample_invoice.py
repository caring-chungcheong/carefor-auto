# -*- coding: utf-8 -*-
"""8-1-1 증빙에서 위펀 거래명세서 엑셀 1개를 받아 구조 파악 (스키마 학습용)."""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from playwright.sync_api import sync_playwright

from src.config import Config, config_path
from src.carefor_client import extract_g_pammgno
from .explore_pages import login
from .branch_pages import _goto, CLOSE_MODAL_JS

OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent / "audit_results" / "explore"


def main():
    branch = sys.argv[1] if len(sys.argv) > 1 else "천안점"
    cfg = Config.load(config_path())
    b = next(x for x in cfg.branches if branch in x.name)
    with sync_playwright() as p:
        browser, page = login(p, b.ctmnumb)
        g = extract_g_pammgno(page)
        _goto(page, "welfare", g)
        page.evaluate(CLOSE_MODAL_JS)
        page.wait_for_timeout(800)
        # '거래명세서' 파일 링크 중 첫 번째 클릭 → 다운로드 캡처
        loc = page.locator("text=/위펀_전체_거래명세서.*xlsx/").first
        name = loc.text_content().strip()
        print("다운로드 대상:", name)
        with page.expect_download(timeout=60000) as dl:
            loc.click()
        d = dl.value
        dest = OUT / d.suggested_filename
        d.save_as(str(dest))
        print("저장:", dest)
        browser.close()

    # 구조 파악
    from openpyxl import load_workbook
    wb = load_workbook(dest, read_only=True)
    print("시트:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    for r, row in enumerate(ws.iter_rows(max_row=15, max_col=12, values_only=True), 1):
        vals = ["" if c is None else str(c)[:14] for c in row]
        if any(vals):
            print(f"  {r:>2}:", " | ".join(vals))


if __name__ == "__main__":
    main()
