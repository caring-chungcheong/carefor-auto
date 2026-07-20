# -*- coding: utf-8 -*-
"""욕구사정 총평 ↔ 판단근거 나란히 보기 (눈 대조용).

사례집 20번(최고위험): 판단근거에 쓴 핵심이 총평에도, 총평에 쓴 내용이 판단근거에도
반영돼 있어야 한다. **양방향 일치**를 봐야 한다.

⚠️ 자동 판정은 안 한다(시도했다 폐기). 총평은 요약이라 표현이 달라(치매↔인지저하,
   고혈압을 약 이름으로만 적음 등) 진단명 substring 대조가 767개 중 283개를 "누락"으로
   찍는 오탐 홍수였다 — 진짜 누락과 동의어 차이를 기계가 못 가른다.
   → **총평과 판단근거를 한 화면에 나란히 놓고 사람이 눈으로 대조**하는 게 유일하게 안전·유용.

실행: py -X utf8 -m audit.make_needs_summary_crosscheck [지점키=청주]
출력: 평가준비/<지점>/<지점>_총평판단근거_나란히.xlsx  (개인정보 → 로컬만, repo 밖)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from audit.deskpath import out_dir

RES = Path(__file__).resolve().parent.parent / "audit_results"

# 판단근거 섹션(총평과 나란히 놓을 항목)
SECTIONS = [
    ("질병", "판단근거(주요 질병"),
    ("영양·배설", "판단근거(영양"),
    ("신체", "판단근거(신체"),
    ("인지", "판단근거(인지"),
    ("의사소통", "판단근거(의사소통"),
    ("가족·환경", "판단근거(가족"),
    ("자원", "판단근거(자원"),
]


def _text(rows, needle):
    for r in rows:
        if needle in str(r.get("label", "")):
            return (r.get("text") or "").replace("\u200E", "").strip()
    return ""


def build(branch_key):
    src = next(RES.glob(f"needs_full_*{branch_key}*.json"))
    d = json.loads(src.read_text(encoding="utf-8"))
    cutoff = d.get("cutoff", "")
    out = []
    for p in d["people"]:
        for a in p.get("assess", []):
            rows = a.get("rows") or []
            if not rows:
                continue
            dt = a.get("date", "")
            if cutoff and dt and dt < cutoff:
                continue
            summ = _text(rows, "총평")
            secs = {name: _text(rows, needle) for name, needle in SECTIONS}
            if summ or any(secs.values()):
                out.append((p["name"], dt, summ, secs))
    return out, cutoff


def main():
    key = sys.argv[1] if len(sys.argv) > 1 else "청주"
    data, cutoff = build(key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "총평↔판단근거 나란히"
    heads = ["수급자", "사정일", "총평(요약)"] + [n for n, _ in SECTIONS]
    ws.append(heads)

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="A94464")
    thin = Side(style="thin", color="E0D3D8")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
    # 총평 열 강조색
    ws.cell(1, 3).fill = PatternFill("solid", fgColor="7A2E45")

    for name, dt, summ, secs in sorted(data, key=lambda x: (str(x[0]), str(x[1]))):
        row = [name, dt, summ] + [secs[n] for n, _ in SECTIONS]
        ws.append(row)
        rr = ws.max_row
        for c in ws[rr]:
            c.border = bd
            c.alignment = Alignment(vertical="top", wrap_text=(c.column >= 3))
        ws.row_dimensions[rr].height = 150

    widths = [10, 12, 60] + [40] * len(SECTIONS)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "C2"   # 수급자·사정일 고정

    # 안내 시트
    ws2 = wb.create_sheet("보는 법")
    guide = [
        ["욕구사정 총평 ↔ 판단근거 나란히 — 보는 법"],
        [""],
        ["• 총평(C열)에 쓴 내용이 오른쪽 판단근거 항목들에 실제로 서술돼 있는지,"],
        ["  반대로 각 판단근거의 핵심이 총평에 반영됐는지 한 줄씩 눈으로 대조하세요."],
        ["• 예: 총평에 '치매'라 썼는데 질병 판단근거엔 고혈압·부정맥만 있고 치매 언급이 없으면 → 보완."],
        [""],
        ["※ 자동 '일치/불일치' 판정은 넣지 않았습니다."],
        ["  총평은 요약이라 표현이 달라(치매↔인지저하 등) 기계가 대조하면 오탐이 많습니다."],
        ["  → 이 표는 두 글을 나란히 놓아 사람이 판단하도록 돕는 용도입니다."],
        [""],
        ["※ 개인정보(수급자명·병력)라 이 파일은 로컬에만 두고 개별 전달하세요."],
    ]
    for g in guide:
        ws2.append(g)
    ws2.column_dimensions["A"].width = 90

    out = out_dir(key) / f"{key}_총평판단근거_나란히.xlsx"
    wb.save(out)
    print(f"저장: {out}")
    print(f"사정 {len(data)}건 나란히 · 평가기간 {cutoff}~")
    print("자동 판정 없음 — 총평·판단근거를 나란히 놓아 사람이 눈으로 대조(오탐 방지)")


if __name__ == "__main__":
    main()
